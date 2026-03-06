import os
import re
import json
import base64
import requests
import streamlit as st
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from supabase import create_client, Client
from dotenv import load_dotenv
load_dotenv()
# ===================== CONFIG =====================

st.set_page_config(
    page_title="Student Comment Generator",
    page_icon="📝",
    layout="wide"
)

ADMIN_EMAILS = [
    "jiyakuttan@gmail.com",
    "admin@example.com"
]

OPENROUTER_API_URL = "https://openrouter.ai/api/v1/chat/completions"

# ===================== ENV =====================

OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY")
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")

if not all([OPENROUTER_API_KEY, SUPABASE_URL, SUPABASE_KEY]):
    st.error("❌ Missing environment variables. Please configure OPENROUTER_API_KEY, SUPABASE_URL, and SUPABASE_SERVICE_KEY")
    st.stop()

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ===================== HELPERS =====================

def is_valid_email(email):
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def is_admin(email):
    return email.lower() in [e.lower() for e in ADMIN_EMAILS]

def get_base64(file):
    if os.path.exists(file):
        with open(file, "rb") as f:
            return base64.b64encode(f.read()).decode()
    return None

def set_background(png):
    b64 = get_base64(png)
    if b64:
        st.markdown(
            f"""
            <style>
            .stApp {{
                background-image: url("data:image/png;base64,{b64}");
                background-size: cover;
            }}
            </style>
            """,
            unsafe_allow_html=True
        )

# ===================== SUPABASE FUNCTIONS =====================

def upsert_user(email, username):
    """Create or update user in Supabase"""
    now = datetime.utcnow().isoformat()
    try:
        supabase.table("users").upsert({
            "email": email,
            "username": username,
            "is_admin": is_admin(email),
            "login_count": 1,
            "first_login": now,
            "last_login": now
        }, on_conflict="email").execute()
        
        supabase.rpc("increment_login", {"user_email": email}).execute()
    except Exception as e:
        st.error(f"Error updating user: {str(e)}")

def increment_comments(email):
    """Increment comment count for user"""
    try:
        supabase.rpc("increment_comments", {"user_email": email}).execute()
    except Exception as e:
        st.error(f"Error incrementing comments: {str(e)}")

def fetch_all_users():
    """Fetch all users from Supabase"""
    try:
        response = supabase.table("users").select("*").execute()
        return response.data
    except Exception as e:
        st.error(f"Error fetching users: {str(e)}")
        return []

def save_comment(data):
    """Save comment to Supabase"""
    try:
        supabase.table("comments").insert(data).execute()
    except Exception as e:
        st.error(f"Error saving comment: {str(e)}")

# ===================== API FUNCTIONS =====================

def query_mistral(prompt):
    """Query the Mistral API via OpenRouter"""
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": "mistralai/mixtral-8x7b-instruct",
        "messages": [{"role": "user", "content": prompt}]
    }
    
    try:
        response = requests.post(OPENROUTER_API_URL, json=payload, headers=headers)
        data = response.json()
        
        if "error" in data:
            return f"API Error: {data['error']}"
        if "choices" not in data or not data["choices"]:
            return f"Error: Invalid API response format"
        
        return data["choices"][0]["message"]["content"]
    except Exception as e:
        return f"Error processing response: {str(e)}"

# ===================== SESSION STATE =====================

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.email = ""
    st.session_state.username = ""

if "workbook" not in st.session_state:
    st.session_state.workbook = Workbook()

# ===================== STYLING =====================

set_background("back6.png")

if os.path.exists('./stylesheet.css'):
    with open('./stylesheet.css') as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

# ===================== LOGIN =====================

if not st.session_state.logged_in:
    st.markdown('<h1 style="text-align: center;">Student Comment Generator</h1>', unsafe_allow_html=True)
    st.markdown('<h3 style="text-align: center;">Please login to continue</h3>', unsafe_allow_html=True)
    
    col_login1, col_login2, col_login3 = st.columns([1, 2, 1])
    with col_login2:
        username = st.text_input("Enter your name:", key="username_input")
        email = st.text_input("Enter your email:", key="email_input")
        login_button = st.button("Login", use_container_width=True)
        
        if login_button:
            if not username.strip():
                st.error("❌ Please enter your name")
            elif not email.strip():
                st.error("❌ Please enter your email")
            elif not is_valid_email(email.strip()):
                st.error("❌ Please enter a valid email address")
            else:
                st.session_state.logged_in = True
                st.session_state.username = username.strip()
                st.session_state.email = email.strip().lower()
                upsert_user(st.session_state.email, st.session_state.username)
                st.success("✅ Login successful!")
                st.rerun()
    
    st.stop()

# ===================== SIDEBAR =====================

user_is_admin = is_admin(st.session_state.email)
if user_is_admin:
    with st.sidebar:
        st.markdown(f"### Welcome, {st.session_state.username}! 👋")
        st.markdown(f"**Email:** {st.session_state.email}")
        
        if user_is_admin:
            st.success("🔑 Admin Access")
        
        # Fetch user stats from Supabase
        try:
            user_data = supabase.table("users").select("*").eq("email", st.session_state.email).execute()
            if user_data.data:
                user_stats = user_data.data[0]
                st.markdown("---")
                st.markdown("### 📊 Your Statistics")
                st.metric("Total Logins", user_stats.get("login_count", 0))
                st.metric("Comments Generated", user_stats.get("comments_generated", 0))
                
                if user_stats.get("last_login"):
                    st.info(f"**Last Login:** {user_stats['last_login']}")
        except Exception as e:
            st.warning("Could not load user statistics")
        
        st.markdown("---")
        if st.button("Logout", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.username = ""
            st.session_state.email = ""
            st.rerun()

# ===================== ADMIN DASHBOARD =====================

if user_is_admin:
    st.markdown('<h1 style="text-align: center;">🔐 Admin Dashboard</h1>', unsafe_allow_html=True)
    st.markdown('<h3 style="text-align: center;">User Statistics Overview</h3>', unsafe_allow_html=True)
    
    users = fetch_all_users()
    
    if not users:
        st.info("No user data available yet.")
    else:
        total_users = len(users)
        total_logins = sum(u.get("login_count", 0) for u in users)
        total_comments = sum(u.get("comments_generated", 0) for u in users)
        
        col1, col2, col3 = st.columns(3)
        col1.metric("Total Users", total_users)
        col2.metric("Total Logins", total_logins)
        col3.metric("Total Comments Generated", total_comments)
        
        if total_users > 0:
            st.metric("Average Comments per User", f"{total_comments / total_users:.1f}")
        
        st.markdown("---")
        st.markdown("### 👥 Detailed User Statistics")
        
        # Sort users by most comments generated
        sorted_users = sorted(users, key=lambda x: x.get('comments_generated', 0), reverse=True)
        
        for user in sorted_users:
            with st.expander(f"{user.get('username', 'Unknown')} ({user.get('email')}) — {user.get('comments_generated', 0)} comments"):
                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**Login Count:** {user.get('login_count', 0)}")
                    st.write(f"**Comments Generated:** {user.get('comments_generated', 0)}")
                with col2:
                    st.write(f"**First Login:** {user.get('first_login', 'N/A')}")
                    st.write(f"**Last Login:** {user.get('last_login', 'N/A')}")
    
    st.markdown(
        """
        <style>
        .footer {
            position: fixed;
            bottom: 0;
            left: 0;
            width: 100%;
            background-color: rgba(255, 255, 255, 0.6);
            color: black;
            text-align: center;
            padding: 10px;
            font-size: 14px;
        }
        </style>
        <div class="footer">
            Designed and developed by Jiya Mary Jerin | For assistance/support reach out to jiyamaryjerin04@gmail.com
        </div>
        """,
        unsafe_allow_html=True
    )
    
    st.stop()

# ===================== USER INTERFACE =====================

workbook = st.session_state.workbook
sheet = workbook.active

if sheet.max_row == 1:
    sheet["A1"] = "Student Comments"
    sheet["A3"] = "Name"
    sheet["B3"] = "Comment"

col1, spacer, col2 = st.columns([1.3, 0.2, 1])

with col1:
    with st.form("comment_form"):
        st.markdown('<p class="title">Student Comment Generator</p>', unsafe_allow_html=True)
        grade = st.text_input("Enter grade:")
        name = st.text_input("Enter name:")
        gender = st.text_input("Enter gender:")
        strength = st.text_input("Enter strengths:")
        weakness = st.text_input("Enter weaknesses:")
        
        col3, col4 = st.columns([1, 1])
        with col3:
            style = st.radio("Select Comment Style:", ["Simple", "Funny", "Formal"])
        with col4:
            size = st.radio("Select Length:", ["50 words", "100 words", "150 words"])
        
        submit = st.form_submit_button("Generate Comment")

with col2:
    if submit:
        if not all([grade, name, gender, strength, weakness]):
            st.error("❌ Please fill in all fields")
        else:
            prompt = (
                f"Give a {style} progress card comment in {size} for a student named {name}. "
                f"The student is a {gender} whose strengths include {strength}. "
                f"Weaknesses include {weakness}. "
                "Make it sound positive. Write in third person. Do not add any emojis."
            )
            
            with st.spinner("Generating comment..."):
                comment_text = query_mistral(prompt)
            
            if not comment_text.startswith("Error") and not comment_text.startswith("API Error"):
                st.success("✅ Comment Generated!")
                st.write(comment_text)
                
                st.markdown("""
                <style>
                .disclaimer {
                    color: black;
                    text-align: center;
                    font-size: 10px;
                    font-style: italic;
                }
                </style>
                <p class="disclaimer">
                    **This is a machine generated response using AI. Please review before use**
                </p>
                """, unsafe_allow_html=True)
                
                # Save to Supabase
                save_comment({
                    "email": st.session_state.email,
                    "student_name": name,
                    "grade": grade,
                    "gender": gender,
                    "strengths": strength,
                    "weaknesses": weakness,
                    "style": style,
                    "size": size,
                    "comment": comment_text,
                    "created_at": datetime.utcnow().isoformat()
                })
                
                increment_comments(st.session_state.email)
                
                # Add to Excel
                next_row = sheet.max_row + 1
                sheet.cell(row=next_row, column=1, value=name)
                sheet.cell(row=next_row, column=2, value=comment_text)
                
                # Download button
                excel_buffer = BytesIO()
                workbook.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.download_button(
                    label="📥 Download Excel File",
                    data=excel_buffer,
                    file_name=f"student_comments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.balloons()
            else:
                st.error(comment_text)

# ===================== FOOTER =====================

st.markdown(
    """
    <style>
    .footer {
        position: fixed;
        bottom: 0;
        left: 0;
        width: 100%;
        background-color: rgba(255, 255, 255, 0.6);
        color: black;
        text-align: center;
        padding: 10px;
        font-size: 14px;
    }
    </style>
    <div class="footer">
        Designed and developed by Jiya Mary Jerin | For assistance/support reach out to jiyamaryjerin04@gmail.com
    </div>
    """,
    unsafe_allow_html=True
)




